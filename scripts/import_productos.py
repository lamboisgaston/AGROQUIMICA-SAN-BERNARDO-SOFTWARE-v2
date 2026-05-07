import sqlite3
import sys
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path(__file__).resolve().parent.parent / 'prisma' / 'dev.db'
EXCEL_PATH = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent / 'precios.xlsx'


HEADER_ALIASES = {
    'nombre': {'nombre', 'producto', 'descripcion', 'descripción'},
    'categoria': {'categoria', 'categoría', 'rubro', 'familia'},
    'marca': {'marca'},
    'unidad': {'unidad', 'presentacion', 'presentación'},
    'stock': {'stock', 'existencia', 'cantidad'},
    'monedaCosto': {'monedacosto', 'monedacompra', 'moneda', 'divisa'},
    'costoBase': {'costobase', 'costocompra', 'costo', 'precio costo'},
    'precioVenta': {'precioventa', 'precio', 'preciofinalpesos', 'precio final pesos'},
    'porcentajeUva': {'porcentajeuva', 'ivaporcentaje', 'iva', 'iva%'},
    'porcentajeFlete': {'porcentajeflete', 'fleteporcentaje', 'flete', 'flete%'},
    'porcentajeGanancia': {'porcentajeganancia', 'gananciaporcentaje', 'margen', 'margen%'},
    'precioUsd': {'preciousd', 'usd', 'precio usd'}
}


def text(v):
    return str(v or '').strip()


def normalize_label(value):
    base = text(value).lower()
    base = unicodedata.normalize('NFD', base)
    base = ''.join(ch for ch in base if unicodedata.category(ch) != 'Mn')
    return ' '.join(base.split())


def canonical_header(value):
    normalized = normalize_label(value)
    compact = normalized.replace(' ', '')
    for canonical, aliases in HEADER_ALIASES.items():
        normalized_aliases = {normalize_label(a) for a in aliases}
        compact_aliases = {a.replace(' ', '') for a in normalized_aliases}
        if normalized in normalized_aliases or compact in compact_aliases:
            return canonical
    return normalized


def parse_number(v, default=0.0):
    if v is None or v == '':
        return default
    s = str(v).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return default


def normalized_key_part(value):
    return ' '.join(text(value).lower().split())


def key(d):
    return '|'.join([
        normalized_key_part(d['nombre']),
        normalized_key_part(d['categoria']),
        normalized_key_part(d['marca']),
        normalized_key_part(d['unidad'])
    ])


def normalize(row):
    moneda = text(row.get('monedaCosto')).upper()
    moneda = 'USD' if moneda == 'USD' else 'ARS'
    costo = parse_number(row.get('costoBase'))
    precio_final = parse_number(row.get('precioVenta'))
    precio_usd_raw = row.get('precioUsd')
    precio_usd = parse_number(precio_usd_raw) if precio_usd_raw not in (None, '') else (costo if moneda == 'USD' else None)

    return {
        'nombre': text(row.get('nombre')),
        'categoria': text(row.get('categoria')),
        'marca': text(row.get('marca')),
        'unidad': text(row.get('unidad')),
        'stock': int(parse_number(row.get('stock'), 0)),
        'monedaCosto': moneda,
        'costoBase': costo,
        'precioVenta': precio_final,
        'porcentajeUva': parse_number(row.get('porcentajeUva')),
        'porcentajeFlete': parse_number(row.get('porcentajeFlete')),
        'porcentajeGanancia': parse_number(row.get('porcentajeGanancia')),
        'precioFinalPesos': precio_final,
        'precioUsd': precio_usd,
    }


def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_headers = [text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [canonical_header(h) for h in raw_headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute('SELECT * FROM Producto')
    existentes = {key(dict(r)): dict(r) for r in cur.fetchall()}

    creados = actualizados = ignorados = 0
    for raw in rows:
        p = normalize(raw)
        if not p['nombre'] or not p['categoria']:
            ignorados += 1
            continue
        k = key(p)
        if k in existentes:
            cur.execute('''
                UPDATE Producto SET nombre=?, categoria=?, marca=?, unidad=?, stock=?, monedaCosto=?, costoBase=?, precioVenta=?, porcentajeUva=?, porcentajeFlete=?, porcentajeGanancia=?, precioFinalPesos=?, precioUsd=? WHERE id=?
            ''', (p['nombre'], p['categoria'], p['marca'], p['unidad'], p['stock'], p['monedaCosto'], p['costoBase'], p['precioVenta'], p['porcentajeUva'], p['porcentajeFlete'], p['porcentajeGanancia'], p['precioFinalPesos'], p['precioUsd'], existentes[k]['id']))
            actualizados += 1
        else:
            cur.execute('''
                INSERT INTO Producto (nombre,categoria,marca,unidad,stock,monedaCosto,costoBase,precioVenta,porcentajeUva,porcentajeFlete,porcentajeGanancia,precioFinalPesos,precioUsd)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (p['nombre'], p['categoria'], p['marca'], p['unidad'], p['stock'], p['monedaCosto'], p['costoBase'], p['precioVenta'], p['porcentajeUva'], p['porcentajeFlete'], p['porcentajeGanancia'], p['precioFinalPesos'], p['precioUsd']))
            p_with_id = dict(p)
            p_with_id['id'] = cur.lastrowid
            existentes[k] = p_with_id
            creados += 1
    conn.commit()
    conn.close()
    print(f'Importación finalizada. Creados: {creados}. Actualizados: {actualizados}. Ignorados: {ignorados}.')


if __name__ == '__main__':
    main()
